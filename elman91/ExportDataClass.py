#!/usr/bin/python2.7
# -*- coding: utf-8 -*-
import time
import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font
from pylab import *
import os
import sys
import random
import re
import os

import openpyxl
import time
from openpyxl.styles import colors
from openpyxl.styles import Font
import networkx as nx
import json, simplejson
from json import dumps, loads, JSONEncoder, JSONDecoder
import matplotlib.pyplot as plt
from openpyxl.styles import  Fill
from openpyxl.cell import Cell
import operator
#from path import path

from networkx.readwrite import json_graph

import time
from openpyxl.styles import colors
from openpyxl.styles import Font

from Tkinter import *
import tkMessageBox

from sys import platform as _platform
import platform


import numpy as np
#from mayavi import mlab
#from matplotlib import pyplot as plt
from scipy.cluster.hierarchy import dendrogram, linkage
import pydot
import csv

from numbers import Number


class ExportData:
    def __init__(self,nameFile=""):
        self.Folder = os.path.dirname(__file__)
        self.path = self.create_the_output_as_ExcelFile(nameFile)
        print "Export of data in  : ", self.path, "\n"

    def set_path(self, _path):
        self.path = _path

    def set_folder(self, _folder):
        self.Folder = _folder

        # ------------------------------------------------------------------------------------------

    def get_path(self):
        return self.path

    def get_folder(self):
        return self.Folder

    # ------------------------------------------------------------------------------------------
    def name_output_ExcelFile(self, FileName, FileType):

        element= self.Folder
        if os.path.isfile(element) == True:
            folderPath = os.path.dirname(element)

        elif (os.path.isdir(element)) == True:
            folderPath = element

        myExcelFile = time.strftime("%Y%m%d") + "_" + FileName + "." + FileType

        return folderPath + "/" + myExcelFile


    def create_the_output_as_ExcelFile(self,nameFile):

        FilePath = self.Folder + "/" +time.strftime("%Y%m%d") + "_DataExtraction"+nameFile+".xlsx"
        if (os.path.exists(FilePath)==True and  os.path.isfile(FilePath)==True):
            os.remove(FilePath)

        return FilePath

    def export(self, data, parameters):

        ExcelFileName = self.path
        if ExcelFileName == "":
            return 0

        print "ExcelFileName : ", self.path, "\n"

        workbook = openpyxl.Workbook()

        Configworksheet = workbook.create_sheet(0)
        Configworksheet.title = "Configuration"

        Configworksheet['A1'] = "List of all the parameters :  "
        Configworksheet['A3'] = "List of Number of sequences : " + str(parameters[0])
        Configworksheet['A5'] = "List of size of sequences : " + str(parameters[1])
        Configworksheet['A7'] = "List of Window of average computation: : " + str(parameters[2])
        Configworksheet['A9'] = "List of Number of neurons in the hidden layer: " + str(parameters[3])
        Configworksheet['A11'] = "List of Learning rate :  : " + str(parameters[4])
        Configworksheet['A13'] = "List of Momentum :" + str(parameters[5])

        AnalysisWorksheet = workbook.create_sheet()
        AnalysisWorksheet.title = "Analyse"

        NumColumn = 1
        AnalysisWorksheet.cell(row=1, column=NumColumn).value = "Essais"
        AnalysisWorksheet.cell(row=1, column=NumColumn).font = Font(color=colors.RED, bold=True)
        NumColumn+=1

        AnalysisWorksheet.cell(row=1, column=NumColumn).value = "Nombre de Séquence"
        AnalysisWorksheet.cell(row=1, column=NumColumn).font = Font(color=colors.RED, bold=True)
        NumColumn += 1

        AnalysisWorksheet.cell(row=1, column=NumColumn).value = "Taille des Séquences"
        AnalysisWorksheet.cell(row=1, column=NumColumn).font = Font(color=colors.RED, bold=True)
        NumColumn += 1

        AnalysisWorksheet.cell(row=1, column=NumColumn).value = "Fenetre de lissage des résultats"
        AnalysisWorksheet.cell(row=1, column=NumColumn).font = Font(color=colors.RED, bold=True)
        NumColumn += 1

        AnalysisWorksheet.cell(row=1, column=NumColumn).value = "Longeur moyenne des séquences"
        AnalysisWorksheet.cell(row=1, column=NumColumn).font = Font(color=colors.RED, bold=True)
        NumColumn += 1

        AnalysisWorksheet.cell(row=1, column=NumColumn).value = "Ecart types des séquences"
        AnalysisWorksheet.cell(row=1, column=NumColumn).font = Font(color=colors.RED, bold=True)
        NumColumn += 1

        AnalysisWorksheet.cell(row=1, column=NumColumn).value = "Nombre de samples"
        AnalysisWorksheet.cell(row=1, column=NumColumn).font = Font(color=colors.RED, bold=True)
        NumColumn += 1

        AnalysisWorksheet.cell(row=1, column=NumColumn).value = "Nombre d'unités de la couche cachée"
        AnalysisWorksheet.cell(row=1, column=NumColumn).font = Font(color=colors.RED, bold=True)
        NumColumn += 1

        AnalysisWorksheet.cell(row=1, column=NumColumn).value = "Learning rate"
        AnalysisWorksheet.cell(row=1, column=NumColumn).font = Font(color=colors.RED, bold=True)
        NumColumn += 1

        AnalysisWorksheet.cell(row=1, column=NumColumn).value = "Momentum "
        AnalysisWorksheet.cell(row=1, column=NumColumn).font = Font(color=colors.RED, bold=True)
        NumColumn += 1

        AnalysisWorksheet.cell(row=1, column=NumColumn).value = "Pourcentage de séquences acceptées"
        AnalysisWorksheet.cell(row=1, column=NumColumn).font = Font(color=colors.RED, bold=True)
        NumColumn += 1

        AnalysisWorksheet.cell(row=1, column=NumColumn).value = "Pourcentage de séquences rejetées"
        AnalysisWorksheet.cell(row=1, column=NumColumn).font = Font(color=colors.RED, bold=True)
        #NumColumn += 1


        NumRow = 2
        for key, value in data.iteritems():
            NumColumn = 1
            """
            data[i] = {'NumberOfSequences': NumberOfSequences, 'window': window,
                       'nbNeuronHiddenLayer': nbNeuronHiddenLayer, 'learningRate': learningRate, 'momentum': momentum}
            data[i]["meanSeq"] = mean(LenSequences)
            data[i]["pstdevSeq"] = pstdev(LenSequences)
            data[i]["nbsamples"] = nbSamples
            data[i]["nbAcceptedSeq"] = (GrammaticalSequencesPercent * 100)
            data[i]["nbRejectedSeq"] = (NonGrammaticalSequencesPercent * 100)
            """

            AnalysisWorksheet.cell(row=NumRow, column=NumColumn).value = int(key)
            NumColumn+=1
            AnalysisWorksheet.cell(row=NumRow, column=NumColumn).value = data[key]["NumberOfSequences"]
            NumColumn +=1
            AnalysisWorksheet.cell(row=NumRow, column=NumColumn).value = data[key]["sizeOfSequence"]
            NumColumn += 1
            AnalysisWorksheet.cell(row=NumRow, column=NumColumn).value = data[key]["window"]
            NumColumn += 1
            AnalysisWorksheet.cell(row=NumRow, column=NumColumn).value = data[key]["meanSeq"]
            NumColumn += 1
            AnalysisWorksheet.cell(row=NumRow, column=NumColumn).value = data[key]["pstdevSeq"]
            NumColumn += 1
            AnalysisWorksheet.cell(row=NumRow, column=NumColumn).value = data[key]["nbsamples"]
            NumColumn += 1
            AnalysisWorksheet.cell(row=NumRow, column=NumColumn).value = data[key]["nbNeuronHiddenLayer"]
            NumColumn += 1
            AnalysisWorksheet.cell(row=NumRow, column=NumColumn).value = data[key]["learningRate"]
            NumColumn += 1
            AnalysisWorksheet.cell(row=NumRow, column=NumColumn).value = data[key]["momentum"]
            NumColumn += 1
            AnalysisWorksheet.cell(row=NumRow, column=NumColumn).value = data[key]["nbAcceptedSeq"]
            NumColumn += 1
            AnalysisWorksheet.cell(row=NumRow, column=NumColumn).value = data[key]["nbRejectedSeq"]
            NumColumn += 1

            NumRow +=1


        workbook.save(ExcelFileName)

